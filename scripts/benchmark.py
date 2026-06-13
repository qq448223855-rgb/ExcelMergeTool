#!/usr/bin/env python3
import argparse
import sys
import tempfile
import time
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from excel_merge_tool import build_merged_workbook, discover_excel_files


def main():
    parser = argparse.ArgumentParser(
        description="Benchmark Excel Merge Tool with a folder of workbooks."
    )
    parser.add_argument("folder", help="Folder containing .xlsx or .xlsm files")
    parser.add_argument(
        "--output",
        help="Optional result path; a temporary file is used by default",
    )
    args = parser.parse_args()

    files = discover_excel_files(args.folder)
    if not files:
        raise SystemExit("No Excel files found.")

    temporary_directory = None
    if args.output:
        output_file = Path(args.output).resolve()
    else:
        temporary_directory = tempfile.TemporaryDirectory()
        output_file = Path(temporary_directory.name) / "benchmark-result.xlsx"

    start = time.perf_counter()
    build_merged_workbook(files, output_file)
    elapsed = time.perf_counter() - start

    workbook = load_workbook(output_file, read_only=True, data_only=False)
    worksheet = workbook.active
    row_count = 0
    column_count = 0
    for row in worksheet.iter_rows():
        row_count += 1
        column_count = max(column_count, len(row))
    print(f"files={len(files)}")
    print(f"rows={row_count}")
    print(f"columns={column_count}")
    print(f"elapsed_seconds={elapsed:.3f}")
    print(f"output_mb={output_file.stat().st_size / 1024 / 1024:.1f}")
    workbook.close()

    if temporary_directory is not None:
        temporary_directory.cleanup()


if __name__ == "__main__":
    main()
