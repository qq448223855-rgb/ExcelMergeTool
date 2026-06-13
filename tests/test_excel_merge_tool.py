import os
import tempfile
import unittest
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from excel_merge_tool import (
    build_merged_workbook,
    discover_excel_files,
    get_file_info,
    get_workbook_metadata,
)


class ExcelMergeToolTests(unittest.TestCase):
    def setUp(self):
        self.temp_directory = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_directory.name)

    def tearDown(self):
        self.temp_directory.cleanup()

    def create_workbook(self, filename, include_merge=False):
        path = self.root / filename
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.column_dimensions["A"].width = 22

        worksheet.append(["名称", "金额", "文本公式", "备注", None])
        worksheet.append([filename, 10, None, "合并内容", None])
        worksheet["B2"] = "=A2"
        worksheet["C2"] = "=保持文本"
        worksheet["C2"].data_type = "s"

        thin = Side(style="thin", color="000000")
        worksheet["A2"].font = Font(bold=True, color="FFFFFF")
        worksheet["A2"].fill = PatternFill("solid", fgColor="336699")
        worksheet["A2"].border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )
        worksheet["B2"].number_format = "#,##0.00"

        if include_merge:
            worksheet.merge_cells("D2:E2")

        workbook.save(path)
        workbook.close()
        return path

    def corrupt_dimension(self, filename):
        replacement = filename.with_suffix(".broken.xlsx")
        with ZipFile(filename) as source, ZipFile(
            replacement,
            "w",
            ZIP_DEFLATED,
        ) as target:
            for item in source.infolist():
                data = source.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    dimension_start = data.find(b"<dimension ")
                    dimension_end = data.find(b"/>", dimension_start)
                    data = (
                        data[:dimension_start]
                        + b'<dimension ref="A1"/>'
                        + data[dimension_end + 2 :]
                    )
                target.writestr(item, data)
        return replacement

    def test_metadata_reads_rows_from_broken_dimension(self):
        source = self.create_workbook("metadata.xlsx")
        broken = self.corrupt_dimension(source)

        metadata = get_workbook_metadata(broken)
        file_info = get_file_info(broken)

        self.assertEqual(metadata.row_count, 2)
        self.assertEqual(file_info["rows"], 2)
        self.assertIn((1, 1, 22.0), metadata.column_widths)

    def test_streaming_merge_preserves_values_formulas_and_styles(self):
        first = self.create_workbook("001.xlsx")
        second = self.create_workbook("002.xlsx")
        output = self.root / "stream-result.xlsx"

        build_merged_workbook([first, second], output, skip_rows=1)

        workbook = load_workbook(output, data_only=False)
        worksheet = workbook.active
        try:
            self.assertEqual(worksheet.max_row, 3)
            self.assertEqual(worksheet["A2"].value, "001.xlsx")
            self.assertEqual(worksheet["A3"].value, "002.xlsx")
            self.assertEqual(worksheet["B2"].value, "=A2")
            self.assertEqual(worksheet["B3"].value, "=A3")
            self.assertEqual(worksheet["C3"].value, "=保持文本")
            self.assertEqual(worksheet["C3"].data_type, "s")
            self.assertTrue(worksheet["A3"].font.bold)
            self.assertEqual(worksheet["A3"].fill.fgColor.rgb, "00336699")
            self.assertEqual(worksheet["B3"].number_format, "#,##0.00")
            self.assertEqual(worksheet.column_dimensions["A"].width, 22.0)
            self.assertEqual(len(worksheet.merged_cells.ranges), 0)
        finally:
            workbook.close()

    def test_merged_cells_trigger_compatible_output(self):
        first = self.create_workbook("001.xlsx")
        second = self.create_workbook("002.xlsx", include_merge=True)
        output = self.root / "merged-result.xlsx"

        build_merged_workbook(
            [first, second],
            output,
            skip_rows=1,
            keep_merged_cells=True,
        )

        workbook = load_workbook(output, data_only=False)
        worksheet = workbook.active
        try:
            self.assertIn("D3:E3", worksheet.merged_cells)
            self.assertEqual(worksheet["D3"].value, "合并内容")
            self.assertEqual(worksheet["A3"].value, "002.xlsx")
        finally:
            workbook.close()

    def test_merged_cells_can_be_disabled_for_streaming_output(self):
        source = self.create_workbook("merged.xlsx", include_merge=True)
        output = self.root / "unmerged-result.xlsx"

        build_merged_workbook(
            [source],
            output,
            keep_merged_cells=False,
        )

        workbook = load_workbook(output)
        try:
            self.assertEqual(len(workbook.active.merged_cells.ranges), 0)
            self.assertEqual(workbook.active["D2"].value, "合并内容")
        finally:
            workbook.close()

    def test_discover_excel_files_is_sorted_and_ignores_temporary_files(self):
        nested = self.root / "nested"
        nested.mkdir()
        self.create_workbook("B.xlsx")
        self.create_workbook("a.xlsx")
        temporary = self.root / "~$ignored.xlsx"
        temporary.write_bytes(b"temporary")
        hidden = nested / ".hidden.xlsx"
        hidden.write_bytes(b"hidden")

        discovered = discover_excel_files(self.root)

        self.assertEqual(
            [Path(filename).name for filename in discovered],
            ["a.xlsx", "B.xlsx"],
        )


if __name__ == "__main__":
    unittest.main()
